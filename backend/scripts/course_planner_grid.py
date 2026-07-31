#!/usr/bin/env python3
"""
Course Planner grid reader / writer
===================================

Backs the in-page editable preview of a generated course planner.

    dump  <xlsx>   -> prints the sheet as JSON on stdout (cells + merges)
    apply <xlsx>   -> reads {"edits": [{"r":1,"c":2,"v":"..."}]} on stdin and
                      writes those cells back into the SAME workbook, in place

"apply" edits only the cell values that changed, so every bit of template
formatting (fonts, fills, merges, column widths, the holiday banners stamped by
generate_course_planner.py) survives untouched.

Requires: openpyxl (vendored under ./vendor)
"""

import datetime
import json
import os
import sys

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "vendor"))

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required.  Install it with:  pip install openpyxl")


# Formats accepted when a user retypes a date cell in the browser.
DATE_INPUT_FORMATS = ("%Y-%m-%d", "%d-%b-%Y", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y")


def merge_map(ws):
    """(r,c) -> (anchor_row, anchor_col) for every merged cell."""
    covered = {}
    for m in ws.merged_cells.ranges:
        for r in range(m.min_row, m.max_row + 1):
            for c in range(m.min_col, m.max_col + 1):
                covered[(r, c)] = (m.min_row, m.min_col)
    return covered


def span_map(ws):
    """(anchor_row, anchor_col) -> (rowspan, colspan)."""
    spans = {}
    for m in ws.merged_cells.ranges:
        spans[(m.min_row, m.min_col)] = (
            m.max_row - m.min_row + 1,
            m.max_col - m.min_col + 1,
        )
    return spans


def last_content_row(ws):
    """Last row holding anything. Templates carry a long tail of blank-but-styled
    rows (the DV planner reports 249 rows for 49 of content); sending those would
    bloat the payload and the rendered grid for no gain."""
    last = 0
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v is not None and str(v).strip() != "":
                last = r
                break
    return last or ws.max_row


def dump(path):
    """Serialise the active sheet for the browser grid."""
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    covered, spans = merge_map(ws), span_map(ws)
    # keep any merge that reaches past the last populated row visible
    last_row = max([last_content_row(ws)] + [m.max_row for m in ws.merged_cells.ranges])
    last_row = min(last_row, ws.max_row)

    rows = []
    for r in range(1, last_row + 1):
        row = []
        for c in range(1, ws.max_column + 1):
            anchor = covered.get((r, c))
            if anchor is not None and anchor != (r, c):
                row.append(None)          # swallowed by a merge -> not rendered
                continue

            v = ws.cell(r, c).value
            kind = "text"
            if isinstance(v, datetime.datetime):
                v = v.date()
            if isinstance(v, datetime.date):
                kind = "date"
                v = v.strftime("%d-%b-%Y")
            elif v is None:
                v = ""
            else:
                v = str(v)

            cell = {"r": r, "c": c, "v": v, "t": kind}
            rs, cs = spans.get((r, c), (1, 1))
            if rs > 1:
                cell["rs"] = rs
            if cs > 1:
                cell["cs"] = cs
            row.append(cell)
        rows.append(row)

    return {
        "sheet": ws.title,
        "maxRow": last_row,
        "maxCol": ws.max_column,
        "rows": rows,
    }


def coerce(cell, text):
    """Turn the edited text back into the cell's original type where sensible,
    so a date stays a real date (and keeps its number format) rather than
    degrading into a string."""
    text = (text or "").replace("\r\n", "\n").strip()
    if text == "":
        return None

    if isinstance(cell.value, (datetime.date, datetime.datetime)):
        for fmt in DATE_INPUT_FORMATS:
            try:
                return datetime.datetime.strptime(text, fmt).date()
            except ValueError:
                continue
        return text                       # not a date any more -> keep the text

    if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
        try:
            return int(text) if text.lstrip("-").isdigit() else float(text)
        except ValueError:
            return text

    return text


def same(a, b):
    """Value equality that ignores the date/datetime split -- openpyxl hands
    back a datetime for a cell we wrote a date into, which would otherwise make
    an unchanged date look like an edit on every save."""
    if isinstance(a, datetime.datetime):
        a = a.date()
    if isinstance(b, datetime.datetime):
        b = b.date()
    return a == b


def apply(path, payload):
    """Write the edited cells back into the workbook, in place."""
    edits = payload.get("edits") or []
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    covered = merge_map(ws)

    changed = 0
    for e in edits:
        try:
            r, c = int(e["r"]), int(e["c"])
        except (KeyError, TypeError, ValueError):
            continue
        if r < 1 or c < 1 or r > ws.max_row or c > ws.max_column:
            continue
        # never write into a cell swallowed by a merge -- openpyxl would raise
        anchor = covered.get((r, c))
        if anchor is not None and anchor != (r, c):
            r, c = anchor

        cell = ws.cell(r, c)
        new = coerce(cell, e.get("v"))
        if not same(new, cell.value):
            cell.value = new
            changed += 1

    wb.save(path)
    return {"ok": True, "saved": changed}


def main():
    if len(sys.argv) < 3:
        sys.exit("usage: course_planner_grid.py <dump|apply> <xlsx>")
    mode, path = sys.argv[1], sys.argv[2]
    if not os.path.exists(path):
        sys.exit(f"workbook not found: {path}")

    if mode == "dump":
        print(json.dumps(dump(path)))
    elif mode == "apply":
        payload = json.load(sys.stdin)
        print(json.dumps(apply(path, payload)))
    else:
        sys.exit(f"unknown mode: {mode}")


if __name__ == "__main__":
    main()
