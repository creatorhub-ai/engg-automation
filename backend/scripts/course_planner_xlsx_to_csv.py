#!/usr/bin/env python3
"""
Course Planner : Generated Excel format  ->  Converted CSV
==========================================================

Flattens a course-planner workbook (the weekly calendar grid produced by the
Course Planner Generator) into the parsing-friendly CSV layout the training
system ingests -- one row per session.

This is the headless, self-configuring version of the original
`course_planner_xlsx_to_csv.py`: instead of hard-coding BATCH_NO / DOMAIN /
session times, it reads them back out of the workbook itself:

    * batch_no   from the "Batch No: <x>" header cell (falls back to the
                 "... Course (<x>)" title)
    * domain     from the "Domain: <x>" header cell
    * session    times from the "Session1 : <t>" lines in the A2 header block
                 (falls back to the "Module Name - S1\\n<t>" column headers)
    * remarks    a "Holiday - <name>" note stamped by the generator is carried
                 into the CSV's remarks column

Output columns (exact order)
----------------------------
    batch_no, domain, mode, week_no, date, start_time, end_time,
    module_name, topic_name, trainer_name, trainer_email, topic_status,
    remarks, batch_type, actual_date, date_difference,
    date_changed_by, date_changed_at

Usage
-----
    python course_planner_xlsx_to_csv.py in.xlsx out.csv

Requires:  pip install openpyxl
"""

import csv
import datetime
import os
import re
import sys

# Prefer the vendored openpyxl/et_xmlfile shipped in ./vendor so the tool works
# on hosts (e.g. Render) with no pip-installed openpyxl. Pure-Python, no build.
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "vendor"))

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required.  Install it with:  pip install openpyxl")


HEADER = [
    "batch_no", "domain", "mode", "week_no", "date", "start_time", "end_time",
    "module_name", "topic_name", "trainer_name", "trainer_email",
    "topic_status", "remarks", "batch_type", "actual_date", "date_difference",
    "date_changed_by", "date_changed_at",
]

MODE = "offline"
STATUS = "Planned"

# session column mapping per domain (1-based): (module_col, topic_col)
#   PD -> S1 C/D , S2 F/G , S3 I (lab column, module==topic)
#   DV -> S1 C/D , S2 G/I (Course / Topic Planned; no lab session)
SESSION_MAPS = {
    "PD": [(3, 4), (6, 7), (9, 9)],
    "DV": [(3, 4), (7, 9)],
}
DEFAULT_MAP = [(3, 4), (6, 7), (9, 9)]   # PD-style fallback (e.g. new domains)

EN_DASH = "–"
DATE_COL = 2


# --------------------------------------------------------------------------- #
# Time parsing:  "7.30AM to 9.00AM" -> ("7:30 AM", "9:00 AM")
# --------------------------------------------------------------------------- #

_T = re.compile(r"(\d{1,2})[.:](\d{2})\s*([APap][Mm])")


def _fmt(h, m, ap):
    return f"{int(h)}:{m} {ap.upper()}"


def parse_range(text):
    """Return (start, end) from a 'start to end' timing string, or ('','')."""
    if not text:
        return "", ""
    hits = _T.findall(str(text))
    if len(hits) >= 2:
        return _fmt(*hits[0]), _fmt(*hits[1])
    if len(hits) == 1:
        return _fmt(*hits[0]), ""
    return "", ""


# --------------------------------------------------------------------------- #
# Merge-aware helpers  (same as the original tool)
# --------------------------------------------------------------------------- #

def build_resolvers(ws):
    anchor, width = {}, {}
    for m in ws.merged_cells.ranges:
        a = ws.cell(m.min_row, m.min_col).value
        for r in range(m.min_row, m.max_row + 1):
            for c in range(m.min_col, m.max_col + 1):
                anchor[(r, c)] = a
                width[(r, c)] = m.max_col - m.min_col
    value = lambda r, c: anchor.get((r, c), ws.cell(r, c).value)   # noqa: E731
    span = lambda r, c: width.get((r, c), 0)                       # noqa: E731
    return value, span


def normalize(text):
    if text is None:
        return ""
    return str(text).replace("–", "-").replace("—", "-").strip()


def split_title(raw):
    full = normalize(raw)
    if raw is not None and EN_DASH in str(raw):
        return str(raw).split(EN_DASH, 1)[0].strip(), full
    return full, full


def trainer_for(module):
    m = module.lower()
    if "soft skill" in m:
        return "Rani", "customer.success@chipedge.com"
    if module == "Lab" or ("assessment" in m and "quiz" not in m):
        return "Akanksha", "akanksha.kumbar@chipedge.com"
    return "Obulesu", "obulesu.b@chipedge.com"


# --------------------------------------------------------------------------- #
# Metadata read-back
# --------------------------------------------------------------------------- #

def read_batch_domain(ws):
    """Pull batch_no, domain and the raw per-session timing strings."""
    batch_no = domain = title = ""
    sessions = ["", "", ""]          # timing strings for S1, S2, S3
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if not isinstance(v, str):
                continue
            s = v.strip()
            low = s.lower()
            if low.startswith("batch no:"):
                batch_no = s.split(":", 1)[1].strip()
            elif low.startswith("domain:"):
                domain = s.split(":", 1)[1].strip()
            if "course" in low and "(" in s and ")" in s and not title:
                title = s
            for i in range(3):
                # stop the timing at the next "SessionN" label (DV keeps S1 & S2
                # on the same line) and at any newline
                m = re.search(rf"session\s*{i + 1}\s*:\s*(.+)", s, re.I)
                if m and not sessions[i]:
                    frag = m.group(1).splitlines()[0]
                    frag = re.split(r"session\s*\d\s*:", frag, flags=re.I)[0]
                    sessions[i] = frag.strip()
    # fallback: batch_no from the title "... Course (PDFT19)"
    if not batch_no and title:
        m = re.search(r"\(([^)]+)\)\s*$", title)
        if m:
            batch_no = m.group(1).strip()
    return batch_no, domain, sessions


def read_times(ws, sessions, session_cols):
    """Resolve (start,end) per session, filling gaps from the
    'Module Name - Sn\\n<timing>' column headers."""
    for i, (mcol, _) in enumerate(session_cols):
        if i < len(sessions) and not sessions[i]:
            hdr = ws.cell(4, mcol).value
            if isinstance(hdr, str) and "\n" in hdr:
                sessions[i] = hdr.split("\n", 1)[1].strip()
    return [parse_range(sessions[i]) if i < len(sessions) else ("", "")
            for i in range(len(session_cols))]


# --------------------------------------------------------------------------- #
# Core conversion
# --------------------------------------------------------------------------- #

def convert(xlsx_path, csv_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    value, span = build_resolvers(ws)
    batch_no, domain, sessions = read_batch_domain(ws)
    session_cols = SESSION_MAPS.get((domain or "").upper(), DEFAULT_MAP)
    times = read_times(ws, sessions, session_cols)

    # remarks column = the generator's trailing "Remarks" header, if any
    # (search the header block: PD labels it on row 4, DV on row 5)
    remarks_col = None
    for hr in range(1, 7):
        for c in range(1, ws.max_column + 1):
            if str(ws.cell(hr, c).value).strip().lower() == "remarks":
                remarks_col = c
                break
        if remarks_col:
            break

    def make_row(week, date_str, start, end, module, topic, remarks=""):
        tname, temail = trainer_for(module)
        return {
            "batch_no": batch_no, "domain": domain, "mode": MODE,
            "week_no": "" if week is None else week,
            "date": date_str, "start_time": start, "end_time": end,
            "module_name": module, "topic_name": topic,
            "trainer_name": tname, "trainer_email": temail,
            "topic_status": STATUS, "remarks": remarks, "batch_type": "",
            "actual_date": "", "date_difference": "",
            "date_changed_by": "", "date_changed_at": "",
        }

    # full-day block spans session1 start .. last available session end
    day_start = times[0][0] or "7:30 AM"
    day_end = next((e for _, e in reversed(times) if e), "1:15 PM")

    rows, seen, week = [], set(), None

    for r in range(1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        if isinstance(a, (int, float)) and not isinstance(a, bool):
            week = int(a)

        b = ws.cell(r, DATE_COL).value
        if not isinstance(b, datetime.datetime):
            continue
        d = b.date()
        date_str = f"{d.month}/{d.day}/{d.year}"
        remarks = ""
        if remarks_col:
            rv = ws.cell(r, remarks_col).value
            if rv:
                remarks = str(rv).strip()

        # full-day blocks: "Final Assessment ..." / "Self Preparation Time"
        s1_raw = value(r, 3)
        if s1_raw and span(r, 3) >= 2 and \
                re.search(r"final assessment|self preparation", str(s1_raw), re.I):
            module, topic = split_title(s1_raw)
            sig = (date_str, day_start, module, topic)
            if sig not in seen:
                seen.add(sig)
                rows.append(make_row(week, date_str, day_start, day_end,
                                     module, topic, remarks))
            continue

        for idx, (mcol, tcol) in enumerate(session_cols):
            start, end = times[idx] if idx < len(times) else ("", "")
            if mcol == tcol:                      # session 3 (lab column)
                text = normalize(value(r, mcol))
                if not text:
                    continue
                low = text.lower()
                if "lab" in low:
                    module = "Lab"
                elif "quiz" in low:
                    module = "Weekly Quiz"
                else:
                    module = split_title(value(r, mcol))[0]
                topic = text
            else:                                 # sessions 1 & 2
                m_raw, t_raw = value(r, mcol), value(r, tcol)
                if not normalize(m_raw) and not normalize(t_raw):
                    continue
                module = split_title(m_raw if normalize(m_raw) else t_raw)[0]
                topic = normalize(t_raw) if normalize(t_raw) else module

            sig = (date_str, start, module, topic)
            if sig in seen:
                continue
            seen.add(sig)
            rows.append(make_row(week, date_str, start, end, module, topic, remarks))

    with open(csv_path, "w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=HEADER)
        writer.writeheader()
        writer.writerows(rows)

    return len(rows)


def main():
    if len(sys.argv) != 3:
        sys.exit("Usage: python course_planner_xlsx_to_csv.py in.xlsx out.csv")
    n = convert(sys.argv[1], sys.argv[2])
    print(f"Wrote {n} rows to {sys.argv[2]}")


if __name__ == "__main__":
    main()
