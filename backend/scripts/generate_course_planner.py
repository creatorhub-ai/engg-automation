#!/usr/bin/env python3
"""
Course Planner Generator
========================

Fills a domain-specific Course Planner *template* with the batch metadata
(domain, session timings, batch no, lab access timings) chosen in the UI,
stamps working-day dates down the grid starting from a batch start date,
and marks company holidays (from the Holiday list) and weekends.

Offline batches teach Monday-Friday, so two calendar rules shape the grid:

  * A batch that starts mid-week only gets the days left in that week.  The
    template lays its weeks out as blocks of day-rows separated by a
    "Saturday & Sunday - Weekend Break" banner, so the grid is re-flowed: each
    weekend banner is pulled up until no week holds more than five teaching
    days.  A Wednesday start therefore delivers three topics in week 1
    (Wed-Fri) and pushes the rest into week 2, which runs Mon-Fri.  No topic
    is dropped; the course simply finishes a few days later, and the Week No
    column is renumbered to match the new weeks.
  * A company holiday is a whole non-teaching day, so it is banded right
    across the schedule columns -- exactly like the weekend banner -- as
    "HOLIDAY - <name>", instead of being written into a single cell.

The template is picked *dynamically* by domain: the first workbook in the
template directory whose filename contains the domain token AND the words
"Course Planner Template" is used (e.g. domain "PD" -> "PD Course Planner
Template.xlsx", "DV" -> "DV Course Planner Template.xlsx").  Adding a new
"DFT Course Planner Template.xlsx" later works with no code change.

Input  : a JSON object on stdin (see KEYS below)
Output : writes the filled .xlsx to <out_path>; prints a small JSON summary
         to stdout.

JSON keys
---------
    domain        "PD" | "DV" | "DFT" ...
    batch_type    "Offline" | "Online"  (picks the matching template; Online
                  stamps weekend-only dates and fills "(current_weekend)")
    batch_no      free text, e.g. "PDFT19"
    session1      e.g. "7.30AM to 9.00AM"
    session2      e.g. "9.30AM to 11.00AM"
    session3      e.g. "11.15AM to 1.15PM"   (optional; PD only)
    lab_timings   e.g. "11.30AM to 1.30PM"
    start_date    "YYYY-MM-DD"  (optional; defaults handled by caller)
    template_dir  folder holding the *.xlsx templates + holiday file
    holiday_file  path to the holiday workbook (optional)
    out_path      where to write the generated workbook

Requires:  pip install openpyxl
"""

import datetime
import glob
import json
import os
import re
import sys

# Prefer the vendored openpyxl/et_xmlfile shipped in ./vendor so the tool works
# on hosts (e.g. Render) with no pip-installed openpyxl. Pure-Python, no build.
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "vendor"))

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.cell.cell import MergedCell
except ImportError:
    sys.exit("openpyxl is required.  Install it with:  pip install openpyxl")


HOLIDAY_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
HOLIDAY_FONT = Font(bold=True, color="FFC00000")   # dark-red, bold banner text

# columns (1-based) scanned to decide whether a row is a "working-day" row
TOPIC_COLS = range(3, 10)          # C..I  (covers both PD and DV layouts)
DATE_COL = 2                       # column B holds the date
WEEKEND_RE = re.compile(r"weekend|saturday\s*&\s*sunday", re.I)
MONDAY = 0                         # datetime.date.weekday() values
SATURDAY, SUNDAY = 5, 6
TEACHING_DAYS_PER_WEEK = 5         # offline batches teach Monday-Friday


# --------------------------------------------------------------------------- #
# Template discovery
# --------------------------------------------------------------------------- #

def find_template(template_dir, domain, batch_type=""):
    """Pick the template workbook for the chosen domain + batch type.

    Selection rules (all case-insensitive):
      * batch type "Offline" -> filename must START WITH "offline"
      * batch type "Online"  -> filename must START WITH "online"
      * the domain token (PD / DV / DFT) must appear in the filename
      * the words "Course Planner Template" must appear in the filename

    e.g. domain "DV" + "Online"  -> "Online - DV Course Planner Template.xlsx"
         domain "PD" + "Offline" -> "Offline - PD Course Planner Template.xlsx"

    Dynamic so new domains/batch types need no code edit."""
    dom = domain.strip().lower()
    bt = (batch_type or "").strip().lower()          # "online" | "offline" | ""
    candidates = sorted(glob.glob(os.path.join(template_dir, "*.xlsx")))

    def domain_ok(name):
        return re.search(rf"\b{re.escape(dom)}\b", name) is not None

    # 1) Preferred: name starts with the batch type + has domain + is a template.
    if bt:
        for path in candidates:
            name = os.path.basename(path).lower()
            if name.startswith(bt) and "course planner template" in name and domain_ok(name):
                return path

    # 2) Back-compat / no batch type: domain + "course planner template",
    #    still respecting the batch-type prefix when one was given.
    for path in candidates:
        name = os.path.basename(path).lower()
        if "course planner template" in name and domain_ok(name) and (not bt or name.startswith(bt)):
            return path

    # 3) Loosest fallback: any template with the domain token (honour batch type).
    for path in candidates:
        name = os.path.basename(path).lower()
        if "template" in name and dom in name and (not bt or name.startswith(bt)):
            return path

    available = sorted(
        os.path.basename(p) for p in candidates
        if "course planner template" in os.path.basename(p).lower()
    )
    raise FileNotFoundError(
        f"No {batch_type or ''} {domain} course planner template is available. "
        f"Add a workbook named \"{batch_type or '<Offline|Online>'} - {domain} Course Planner "
        f"Template.xlsx\". Templates found: {', '.join(available) or 'none'}")


def _header_text(v):
    """Normalise a header cell for comparison (templates use NBSPs / stray space)."""
    return v.replace("\xa0", " ").strip().lower() if isinstance(v, str) else ""


def find_header_column(ws, data_start, names, default=0):
    """Locate a column by its header label, scanning the rows above the data."""
    limit = min(ws.max_row, max(data_start + 2, 15))
    for r in range(1, limit + 1):
        for c in range(1, ws.max_column + 1):
            if _header_text(ws.cell(r, c).value) in names:
                return c
    return default


def find_date_column(ws, data_start):
    """Locate the 'Date' column by scanning the header rows above the data.
    Templates differ: Offline/Online-PD keep Date in column B, while Online
    DV/DFT put it in column C.  Falls back to column B."""
    return find_header_column(ws, data_start, {"date"}, DATE_COL)


def find_theory_lab_column(ws, data_start):
    """Locate the 'Theory/Lab' column of the Online templates (0 when absent).
    Online batches run Theory on Saturday and Lab on Sunday, so this column --
    not the row order -- is what anchors each row to a weekend day."""
    return find_header_column(ws, data_start, {"theory/lab", "theory / lab"}, 0)


def session_weekday(ws, value, r, tl_col):
    """Weekend day a row is pinned to: SATURDAY for a Theory row, SUNDAY for a
    Lab row, None when the row is neither (e.g. the orientation session)."""
    if not tl_col:
        return None
    t = _header_text(value(r, tl_col))
    if t.startswith("theory"):
        return SATURDAY
    if t.startswith("lab"):
        return SUNDAY
    return None


def on_or_after(d, weekday):
    """First date on/after `d` that falls on `weekday` (Mon=0 .. Sun=6)."""
    return d + datetime.timedelta(days=(weekday - d.weekday()) % 7)


def weekend_of(d):
    """Return (saturday, sunday) for the weekend the date `d` belongs to.
    Online dates are always Sat/Sun; the weekday branch is a safe fallback."""
    wd = d.weekday()                       # Mon=0 .. Sat=5, Sun=6
    if wd == 5:                            # Saturday
        return d, d + datetime.timedelta(days=1)
    if wd == 6:                            # Sunday
        return d - datetime.timedelta(days=1), d
    sat = d + datetime.timedelta(days=(5 - wd))
    return sat, sat + datetime.timedelta(days=1)


def next_weekend_after(d):
    """Sat & Sun of the weekend FOLLOWING the weekend that `d` belongs to."""
    sat, sun = weekend_of(d)
    return sat + datetime.timedelta(days=7), sun + datetime.timedelta(days=7)


def fill_course_break(ws, r, last_stamped, current):
    """Handle the Course Break banner (the "(current_weekend)" placeholder in the
    Online templates).

    The break takes the weekend AFTER the last class weekend, so it forms a real
    gap in the schedule: classes stop, that whole weekend is the break, and the
    course resumes the weekend after it.

    Returns (handled, resume_from) -- resume_from is the date the caller should
    continue stamping from, i.e. the day after the break's Sunday."""
    for c in range(1, ws.max_column + 1):
        tc = target_cell(ws, r, c)
        v = tc.value
        if isinstance(v, str) and "(current_weekend)" in v:
            if last_stamped is not None:
                sat, sun = next_weekend_after(last_stamped)
            else:                              # break before any class row
                sat, sun = weekend_of(current)
            label = f"{sat.strftime('%d-%b-%Y')} & {sun.strftime('%d-%b-%Y')}"
            tc.value = v.replace("(current_weekend)", label)
            return True, sun + datetime.timedelta(days=1)
    return False, None


# --------------------------------------------------------------------------- #
# Holiday list
# --------------------------------------------------------------------------- #

def load_holidays(holiday_file):
    """Return {date: name} for rows whose 'Type of Holiday' == 'Holiday'
    (exact, so 'Restricted Holiday' is treated as a normal working day)."""
    holidays = {}
    if not holiday_file or not os.path.exists(holiday_file):
        return holidays
    wb = openpyxl.load_workbook(holiday_file, data_only=True)
    ws = wb.active
    # locate columns from the header row
    header = {str(ws.cell(1, c).value).strip().lower(): c
              for c in range(1, ws.max_column + 1) if ws.cell(1, c).value}
    date_c = header.get("date", 1)
    name_c = header.get("holiday", 3)
    type_c = header.get("type of holiday", 4)
    for r in range(2, ws.max_row + 1):
        d = ws.cell(r, date_c).value
        t = ws.cell(r, type_c).value
        if isinstance(d, datetime.datetime):
            d = d.date()
        if not isinstance(d, datetime.date):
            continue
        if str(t).strip().lower() == "holiday":
            holidays[d] = str(ws.cell(r, name_c).value or "Holiday").strip()
    return holidays


# --------------------------------------------------------------------------- #
# Merge-aware helpers
# --------------------------------------------------------------------------- #

def build_resolvers(ws):
    """value(r,c) inherits merge anchors; span(r,c) = merge column width."""
    anchor, width = {}, {}
    for m in ws.merged_cells.ranges:
        a = ws.cell(m.min_row, m.min_col).value
        for r in range(m.min_row, m.max_row + 1):
            for c in range(m.min_col, m.max_col + 1):
                anchor[(r, c)] = a
                width[(r, c)] = m.max_col - m.min_col
    value = lambda r, c: anchor.get((r, c), ws.cell(r, c).value)   # noqa: E731
    span = lambda r, c: width.get((r, c), 0)                       # noqa: E731
    return value, span


def target_cell(ws, r, c):
    """Return the writable cell for (r,c): the merge anchor if (r,c) is merged."""
    for m in ws.merged_cells.ranges:
        if m.min_row <= r <= m.max_row and m.min_col <= c <= m.max_col:
            return ws.cell(m.min_row, m.min_col)
    return ws.cell(r, c)


def set_cell(ws, r, c, val):
    """Write to (r,c) honouring merges (write to the merge anchor)."""
    target_cell(ws, r, c).value = val


# --------------------------------------------------------------------------- #
# Metadata fill
# --------------------------------------------------------------------------- #

def fill_metadata(ws, cfg):
    """Fill batch no, domain, session timings and lab timings into the header.
    Robust to layout differences between templates: we edit the label cells
    themselves and replace text placeholders rather than guessing value cells."""
    s1, s2, s3 = cfg.get("session1", ""), cfg.get("session2", ""), cfg.get("session3", "")
    batch, dom, lab = cfg["batch_no"], cfg["domain"], cfg.get("lab_timings", "")

    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if not isinstance(v, str):
                continue
            new = v

            # A1 title placeholder: "... Course (batch_no)"
            new = new.replace("(batch_no)", batch)

            # header timing placeholders (PD template)
            new = new.replace("(session1_timings)", s1)
            new = new.replace("(session2_timinigs)", s2)   # note: template typo
            new = new.replace("(session2_timings)", s2)
            new = new.replace("(session3_timings)", s3)

            # A2 session label lines: "Session1 :" -> "Session1 : 7.30AM to 9.00AM"
            # ([ \t]* not \s*, so we don't swallow the newline between lines)
            new = re.sub(r"(Session\s*1\s*:)[ \t]*", rf"\g<1> {s1}", new)
            new = re.sub(r"(Session\s*2\s*:)[ \t]*", rf"\g<1> {s2}", new)
            if s3:
                new = re.sub(r"(Session\s*3\s*:)[ \t]*", rf"\g<1> {s3}", new)

            # label cells (append the chosen value once)
            low = v.lower()
            if low.strip().rstrip(":") == "domain" or low.startswith("domain:"):
                new = f"Domain: {dom}"
            elif low.strip().rstrip(":") == "batch no" or low.startswith("batch no:"):
                new = f"Batch No: {batch}"
            elif low.startswith("lab access timings") and lab and lab not in v:
                new = f"{v.rstrip()}\n{lab}"

            # DV header: hard-coded timing on 2nd line of "Module Name - S1\n...."
            if s1 and re.match(r"module name\s*-\s*s1", low):
                new = re.sub(r"(Module Name\s*-\s*S1\s*\n).*", rf"\g<1>{s1}", new,
                             flags=re.I | re.S)
            if s2 and re.match(r"module name\s*-\s*s2", low):
                new = re.sub(r"(Module Name\s*-\s*S2\s*\n).*", rf"\g<1>{s2}", new,
                             flags=re.I | re.S)
            if s3 and re.match(r"module name\s*-\s*s3", low):
                new = re.sub(r"(Module Name\s*-\s*S3\s*\n).*", rf"\g<1>{s3}", new,
                             flags=re.I | re.S)

            if new != v:
                cell.value = new


# --------------------------------------------------------------------------- #
# Week re-flow (mid-week batch start)
# --------------------------------------------------------------------------- #

def first_data_row(ws):
    """Row where the schedule starts: the first Week-No cell (col A integer).
    Sits below single- *and* two-row headers (DV has a 'Course/Topic Planned'
    sub-header)."""
    return next(
        (r for r in range(1, ws.max_row + 1)
         if isinstance(ws.cell(r, 1).value, (int, float))
         and not isinstance(ws.cell(r, 1).value, bool)),
        5,
    )


def last_content_row(ws):
    """Last row that holds anything.  ws.max_row over-reports badly on these
    templates (the PD workbook claims 997 rows for 152 rows of schedule), and
    the re-flow must not drag hundreds of blank rows around."""
    for r in range(ws.max_row, 0, -1):
        if any(ws.cell(r, c).value is not None for c in range(1, ws.max_column + 1)):
            return r
    return ws.max_row


def is_weekend_row(value, r, date_col):
    """True for the full-width 'Saturday & Sunday - Weekend Break' banner rows
    that separate one week of the template from the next."""
    v = value(r, date_col)
    return isinstance(v, str) and bool(WEEKEND_RE.search(v))


def weekend_banner_end(ws, data_start, region_end, value, date_col):
    """Last column the weekend banner spans (I on both offline templates).
    Holiday banners copy this span so a holiday reads as a full day off, the
    same way a weekend does.  0 when the template has no weekend row."""
    end = 0
    for r in range(data_start, region_end + 1):
        if not is_weekend_row(value, r, date_col):
            continue
        for m in ws.merged_cells.ranges:
            if m.min_row == r and m.min_col <= date_col <= m.max_col:
                end = max(end, m.max_col)
    return end


def classify_rows(ws, value, span, data_start, region_end, date_col, topic_cols):
    """Tag every schedule row 'weekend', 'day' or 'other' (banners, spacers).
    Only 'day' rows carry a teaching date, and only they count towards the five
    teaching days a week can hold."""
    kind = {}
    for r in range(data_start, region_end + 1):
        if is_weekend_row(value, r, date_col):
            kind[r] = "weekend"
        elif is_day_row(ws, value, span, r, date_col, topic_cols):
            kind[r] = "day"
        else:
            kind[r] = "other"
    return kind


def plan_row_order(kind, data_start, region_end, pad):
    """Re-order the schedule rows so no week holds more than five teaching days.

    `pad` is how many weekdays of the start week are already gone (Mon=0 ..
    Fri=4), so week 1 starts with that many days already spent.  Walking down
    the grid we count teaching days; the moment a sixth would land in the same
    week we pull the next weekend banner up to here, which pushes the surplus
    topics into the following week.  Template weeks that legitimately run short
    (the project and assessment weeks hold 1-4 day-rows) close early on their
    own banner, so their shape is preserved.

    Returns the new ordering of the source row numbers."""
    seq = list(range(data_start, region_end + 1))
    out = []
    day_count = pad
    i = 0
    while i < len(seq):
        r = seq[i]
        k = kind[r]
        if k == "weekend":
            out.append(r)
            day_count = 0
            i += 1
            continue
        if k == "day" and day_count >= TEACHING_DAYS_PER_WEEK:
            nxt = next((x for x in seq[i:] if kind[x] == "weekend"), None)
            if nxt is not None:
                seq.remove(nxt)          # steal it from further down the grid
                out.append(nxt)
                day_count = 0
                continue                 # re-test this day row against a fresh week
        if k == "day":
            day_count += 1
        out.append(r)
        i += 1
    return out


def merges_touching(ws, row_start, row_end):
    """Every merged range that reaches into [row_start..row_end], including the
    ones anchored outside it.  The DV template runs merges straight through the
    boundary (its 'Module Name - S1' cell is F4:F28, and the last Week No cell
    spills past the final row of content), and a range left merged makes every
    cell it covers read-only."""
    return [m for m in list(ws.merged_cells.ranges)
            if m.max_row >= row_start and m.min_row <= row_end]


def _contiguous_runs(nums):
    """[1,2,3,7,8] -> [[1,2,3],[7,8]]"""
    runs, run = [], []
    for n in nums:
        if run and n == run[-1] + 1:
            run.append(n)
        else:
            if run:
                runs.append(run)
            run = [n]
    if run:
        runs.append(run)
    return runs


def apply_row_order(ws, region_start, region_end, order):
    """Rewrite [region_start..region_end] in `order`, carrying values, styles,
    row heights and merges.  openpyxl cannot move rows, so the region is
    snapshotted, unmerged, then written back in the new order.

    A merged range whose rows are no longer adjacent (a week's 'Module Name'
    cell that a re-flowed weekend now cuts in two) is re-created as one merge
    per contiguous run, and the text is repeated so the module name still reads
    on both sides of the break -- except onto a run of rows that were blank in
    the template, which must stay blank (the templates keep a spacer row at the
    end of a week, and text there would read as another teaching day).

    Returns {source row -> new row}."""
    ncols = ws.max_column
    snap = {
        r: [(ws.cell(r, c).value, ws.cell(r, c)._style) for c in range(1, ncols + 1)]
        for r in range(region_start, region_end + 1)
    }
    heights = {r: ws.row_dimensions[r].height
               for r in range(region_start, region_end + 1)}

    touching = merges_touching(ws, region_start, region_end)
    merge_specs = [(m.min_row, m.min_col, m.max_row, m.max_col,
                    ws.cell(m.min_row, m.min_col).value) for m in touching]
    for m in touching:
        ws.unmerge_cells(str(m))

    pos = {s: region_start + i for i, s in enumerate(order)}
    for s, r in pos.items():
        for c, (v, style) in enumerate(snap[s], start=1):
            cell = ws.cell(r, c)
            cell.value = v
            cell._style = style
        ws.row_dimensions[r].height = heights[s]

    blank_src = {r for r, cells in snap.items() if all(v is None for v, _ in cells)}
    for r0, c0, r1, c1, anchor_val in merge_specs:
        # rows outside the re-flowed region did not move, so they map to themselves
        back = {pos.get(r, r): r for r in range(r0, r1 + 1)}
        for run in _contiguous_runs(sorted(back)):
            a, b = run[0], run[-1]
            if a != b or c0 != c1:                     # a 1x1 range is not a merge
                ws.merge_cells(start_row=a, start_column=c0, end_row=b, end_column=c1)
            if back[a] == r0:
                continue                               # anchor already carries the text
            if all(back[x] in blank_src for x in run):
                continue                               # spacer rows stay blank
            ws.cell(a, c0).value = anchor_val
    return pos


def renumber_weeks(ws, data_start, region_end, kind_by_row):
    """Renumber the Week No column so it matches the re-flowed weeks: every
    stretch of rows closed by a weekend banner is one week.  Returns the number
    of weeks written."""
    for m in [m for m in merges_touching(ws, data_start, region_end)
              if m.min_col == 1 and m.max_col == 1]:
        ws.unmerge_cells(str(m))

    # Rows whose column A is swallowed by a wide banner (the DV weekend rows
    # merge A:E / A:I) must not be written to, nor folded into a Week No merge.
    banner_rows = set()
    for m in ws.merged_cells.ranges:
        if m.min_col == 1 and m.max_col > 1:
            banner_rows.update(range(m.min_row, m.max_row + 1))

    blocks, cur = [], []
    for r in range(data_start, region_end + 1):
        if kind_by_row.get(r) == "weekend":
            blocks.append((cur, r))
            cur = []
        else:
            cur.append(r)
    if cur:
        blocks.append((cur, None))

    week = 0
    for rows, weekend_row in blocks:
        if not any(kind_by_row.get(r) == "day" for r in rows):
            continue                       # trailing spacers are not a week
        week += 1
        for r in rows:
            if r not in banner_rows and not isinstance(ws.cell(r, 1), MergedCell):
                ws.cell(r, 1).value = None
        first, last = rows[0], rows[-1]
        if weekend_row is not None and weekend_row not in banner_rows:
            last = weekend_row
        ws.cell(first, 1).value = week
        if last > first:
            ws.merge_cells(start_row=first, start_column=1, end_row=last, end_column=1)
    return week


def reflow_for_start(ws, data_start, date_col, pad):
    """Adjust the grid for a batch that starts mid-week.  Returns
    (weeks, first_week_days): how many weeks the schedule now spans and how
    many teaching days land in week 1."""
    region_end = last_content_row(ws)
    value, span = build_resolvers(ws)
    topic_cols = range(date_col + 1, date_col + 8)
    kind = classify_rows(ws, value, span, data_start, region_end, date_col, topic_cols)

    # Runs even for a Monday start (pad 0): a couple of template blocks carry a
    # sixth day-row, which would otherwise push a teaching day past the weekend
    # banner and cost the batch the rest of that calendar week.
    order = plan_row_order(kind, data_start, region_end, pad)
    if order != list(range(data_start, region_end + 1)):
        pos = apply_row_order(ws, data_start, region_end, order)
        kind = {pos[r]: k for r, k in kind.items()}

    weeks = renumber_weeks(ws, data_start, region_end, kind)
    first_week_days = 0
    for r in range(data_start, region_end + 1):
        if kind.get(r) == "weekend":
            break
        if kind.get(r) == "day":
            first_week_days += 1
    return weeks, first_week_days


# --------------------------------------------------------------------------- #
# Holiday banner
# --------------------------------------------------------------------------- #

def _split_merge_around_row(ws, m, r):
    """Free row `r` from merged range `m`, keeping the parts above and below it
    merged.  Unmerging leaves the text on the original anchor, so the part above
    the holiday still reads; the part below is deliberately left blank rather
    than given a copy -- the row under a Friday holiday is usually the template's
    blank end-of-week spacer, and text there would read as a teaching day."""
    r0, c0, r1, c1 = m.min_row, m.min_col, m.max_row, m.max_col
    ws.unmerge_cells(str(m))
    for a, b in ((r0, r - 1), (r + 1, r1)):
        if b < a or (a == b and c0 == c1):
            continue                       # nothing left that needs a merge
        ws.merge_cells(start_row=a, start_column=c0, end_row=b, end_column=c1)


def mark_holiday_row(ws, r, name, banner_start, banner_end, note_col, last_col):
    """Turn a day row into a full-width holiday banner.

    A company holiday is a whole day off, so it is shown the way the weekend
    break is -- one banner merged straight across the schedule columns --
    rather than a label dropped into the first topic cell.  Any merge reaching
    into this row is split around it first, so the module names above and below
    the holiday survive."""
    for m in list(ws.merged_cells.ranges):
        if m.max_col < banner_start or m.min_col > banner_end:
            continue
        if m.max_row < r or m.min_row > r:
            continue
        _split_merge_around_row(ws, m, r)

    for c in range(banner_start, banner_end + 1):
        ws.cell(r, c).value = None
    if banner_end > banner_start:
        ws.merge_cells(start_row=r, start_column=banner_start,
                       end_row=r, end_column=banner_end)

    banner = ws.cell(r, banner_start)
    banner.value = f"HOLIDAY - {name}"
    banner.font = HOLIDAY_FONT
    banner.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.cell(r, note_col).value = f"Holiday - {name}"      # keep the remark too
    for c in range(1, last_col + 1):
        ws.cell(r, c).fill = HOLIDAY_FILL


# --------------------------------------------------------------------------- #
# Date + holiday fill
# --------------------------------------------------------------------------- #

def is_day_row(ws, value, span, r, date_col, topic_cols):
    """A working-day row: has *its own* topic content and is not a
    weekend/banner row.  Content is checked on RAW cells (ws.cell) not the
    merge-resolved value(), so a merge that leaks a value into a blank spacer
    row (e.g. E10:E11) does not falsely count that spacer as a day."""
    b = value(r, date_col)
    if isinstance(b, str) and WEEKEND_RE.search(b):
        return False
    # full-width banner (e.g. "Foundation Courses") spans many columns -> skip
    if isinstance(b, str) and span(r, date_col) >= 5:
        return False
    for c in topic_cols:
        v = ws.cell(r, c).value                       # RAW, not merge-inherited
        if v is not None and str(v).strip():
            return True
    return False


def fill_dates(ws, start_date, holidays, date_col, online=False):
    """Stamp dates down the day-rows and mark company holidays.

    Offline (default): classes run Mon-Fri.  Weekday dates are stamped down the
    day-rows, weekends are rolled past, and every "Saturday & Sunday - Weekend
    Break" banner closes the week -- the next teaching day is the following
    Monday.  Because the grid has already been re-flowed for the start weekday
    (see reflow_for_start), that keeps each block on exactly one calendar week:
    a mid-week start fills only the days left in week 1 and the rest of the
    course runs Mon-Fri from week 2 on.
    Online: classes run on weekends, so ONLY Saturday/Sunday dates are stamped
    and weekdays are skipped.  A company holiday landing on a weekend does not
    cancel an online session -- the class is scheduled that day regardless.
    Each row is pinned by its Theory/Lab column --
    Theory to Saturday, Lab to Sunday -- rather than by row order, so a
    template that carries an extra row inside the week grid (the DFT
    orientation session) cannot shift Theory onto Sunday.  The Course Break
    banner's "(current_weekend)" placeholder is filled with the Sat & Sun of
    the previous stamped weekend."""
    value, span = build_resolvers(ws)
    # topic columns sit to the right of the Date column; keep the original
    # 7-column scan width so offline behaviour is unchanged (date_col 2 -> C..I).
    topic_cols = range(date_col + 1, date_col + 8)
    current = start_date
    marked = 0
    last_stamped = None
    data_start = first_data_row(ws)
    header_row = data_start - 1
    tl_col = find_theory_lab_column(ws, data_start) if online else 0
    note_col = ws.max_column + 1        # fixed "Remarks" column, computed once
    last_col = note_col
    banner_start = date_col + 1         # first topic/schedule column
    # A holiday is banded across exactly the columns the weekend banner uses,
    # so both read as a full day off.
    banner_end = weekend_banner_end(ws, data_start, last_content_row(ws), value, date_col)
    if banner_end < banner_start:
        banner_end = note_col - 1
    ws.cell(header_row, note_col).value = "Remarks"
    for r in range(data_start, ws.max_row + 1):
        # Online Course Break: the banner takes the weekend after the last class
        # weekend and the schedule resumes past it, so the break is a real gap.
        if online:
            handled, resume_from = fill_course_break(ws, r, last_stamped, current)
            if handled:
                if resume_from is not None:
                    current = resume_from
                    last_stamped = None    # next weekend is fresh, not a break
                continue
        elif is_weekend_row(value, r, date_col):
            # Weekend banner: the week ends here, so teaching resumes on the
            # next Monday even when the week ran short (project/assessment
            # weeks) or started mid-week.
            current = on_or_after(current, MONDAY)
            continue
        if not is_day_row(ws, value, span, r, date_col, topic_cols):
            continue
        if online:
            # Theory -> Saturday, Lab -> Sunday.  Rows that are neither (the
            # orientation session) just take the next weekend day available.
            want = session_weekday(ws, value, r, tl_col)
            if want is not None:
                current = on_or_after(current, want)
            else:
                while current.weekday() < SATURDAY:         # 0-4 = Mon-Fri
                    current += datetime.timedelta(days=1)
        else:
            # topics run Monday-Friday; roll past weekends
            while current.weekday() >= SATURDAY:           # 5=Sat, 6=Sun
                current += datetime.timedelta(days=1)
        set_cell(ws, r, date_col, current)
        ws.cell(r, date_col).number_format = "m/d/yyyy"
        last_stamped = current
        # Online batches already run on the weekend, so a company holiday that
        # falls on one does NOT cancel the session -- the class is scheduled that
        # day anyway.  Offline (weekday) holidays keep the banner behaviour.
        if not online and current in holidays:
            # Company holiday: a non-teaching day.  Drop the topic that fell
            # here and band "HOLIDAY - <name>" right across the schedule
            # columns, so the whole day reads as off exactly like a weekend.
            mark_holiday_row(ws, r, holidays[current], banner_start, banner_end,
                             note_col, last_col)
            marked += 1
        current += datetime.timedelta(days=1)
    return marked


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #

def main():
    cfg = json.load(sys.stdin)
    batch_type = cfg.get("batch_type", "")
    online = batch_type.strip().lower() == "online"
    try:
        template = find_template(cfg["template_dir"], cfg["domain"], batch_type)
    except FileNotFoundError as e:
        # Config problem, not a crash: report the message alone (no traceback)
        # so the API can surface it to the user verbatim.
        sys.exit(str(e))
    holidays = load_holidays(cfg.get("holiday_file"))

    start = cfg.get("start_date")
    if start:
        start_date = datetime.datetime.strptime(start, "%Y-%m-%d").date()
    else:
        start_date = None

    wb = openpyxl.load_workbook(template)
    ws = wb.active

    fill_metadata(ws, cfg)
    holidays_marked = 0
    weeks = 0
    first_week_days = 0
    if start_date:
        # data begins at the first week-number row; use it to locate the Date
        # column (it moves between templates: B for Offline/Online-PD, C for
        # Online DV/DFT).
        data_start = first_data_row(ws)
        date_col = find_date_column(ws, data_start)
        if not online:
            # Offline batches teach Mon-Fri.  A start date on a weekend is not
            # a teaching day, so the batch begins the following Monday.
            if start_date.weekday() >= SATURDAY:
                start_date = on_or_after(start_date, MONDAY)
            # Days of the start week already gone (Mon=0 .. Fri=4): week 1 only
            # gets what is left, and the grid is re-flowed to match.
            pad = start_date.weekday()
            weeks, first_week_days = reflow_for_start(ws, data_start, date_col, pad)
        holidays_marked = fill_dates(ws, start_date, holidays, date_col, online)

    out_path = cfg["out_path"]
    wb.save(out_path)

    print(json.dumps({
        "ok": True,
        "template": os.path.basename(template),
        "out_path": out_path,
        "holidays_marked": holidays_marked,
        "start_date": start,
        "effective_start_date": start_date.isoformat() if start_date else None,
        "start_weekday": start_date.strftime("%A") if start_date else None,
        "weeks": weeks,
        "first_week_days": first_week_days,
        "batch_type": batch_type,
        "online": online,
    }))


if __name__ == "__main__":
    main()
